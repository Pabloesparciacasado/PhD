"""
numerical_greeks.py
-------------------
Model-free option Greeks (delta, gamma) following the numerical methodology in
Bates (2005) "Hedging the Smirk", Table 1.

Pipeline
--------
1.  Filter OPPRCD data for (date, exdate, security_id); obtain S from ForwardPrice.
2.  Fit cubic spline on implied volatility over moneyness (K/S).
3.  Evaluate spline on a uniform moneyness grid.
4.  Reconstruct option prices via Black-Scholes on the uniform grid.
5.  Central finite differences on the uniform-ΔK grid → Bates (1) and (2).

References
----------
Bates, D. S. (2005). Hedging the smirk. Finance Research Letters, 2(4), 195-200.
BKM (2003): Bakshi, Kapadia & Madan (2003), RFS.
Conrad, Dittmar & Ghysels (2013), JF.
"""

from __future__ import annotations

import warnings
from typing import Union

import numpy as np
import pandas as pd
from scipy.interpolate import CubicSpline
from scipy.stats import norm

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from __2_Files.forward_price import ForwardPrice


# ─── Public API ───────────────────────────────────────────────────────────────

def compute_bates_greeks(
    df_opprcd: pd.DataFrame,
    forward_price_source: Union[ForwardPrice, pd.DataFrame],
    date: str,
    exdate: str,
    security_id: int,
    n_grid: int = 100,
) -> pd.DataFrame:
    """
    Compute Bates (2005) model-free delta and gamma for one (date, exdate, security_id).

    Parameters
    ----------
    df_opprcd : pd.DataFrame
        OPPRCD data. Required columns:
        ``date``, ``exdate``, ``security_id``, ``strike_price``, ``best_bid``,
        ``best_offer``, ``cp_flag`` ("C"/"P"), ``impl_volatility``, ``tb_m3``.
    forward_price_source : ForwardPrice | pd.DataFrame
        Either a loaded ForwardPrice instance (with ``.df`` attribute) or a
        DataFrame with columns ``[SecurityID, Date, Expiration, ForwardPrice]``.
    date : str
        Observation date, 'YYYY-MM-DD'.
    exdate : str
        Expiration date, 'YYYY-MM-DD'.
    security_id : int
        OptionMetrics SecurityID of the underlying index.
    n_grid : int
        Number of points on the uniform moneyness grid (default 100).

    Returns
    -------
    pd.DataFrame
        Columns: ``[date, exdate, security_id, moneyness, strike, mid_price,
        iv_spline, bs_price, delta_bates, gamma_bates]``.
        Empty DataFrame (same columns) when data are insufficient.
    """
    # ── 0. Obtain S (forward price) ──────────────────────────────────────────
    S = _get_forward_price(forward_price_source, date, exdate, security_id)
    if S is None or not np.isfinite(S) or S <= 0:
        warnings.warn(
            f"No valid forward price for SecurityID={security_id}, "
            f"date={date}, exdate={exdate}. Returning empty DataFrame.",
            UserWarning, stacklevel=2,
        )
        return _empty_result()

    # ── 1. Filter options ────────────────────────────────────────────────────
    df_filt, r, T = _filter_options(df_opprcd, date, exdate, security_id, S)
    n_obs = len(df_filt) if df_filt is not None else 0
    if n_obs < 4:
        warnings.warn(
            f"Insufficient options after filtering for SecurityID={security_id}, "
            f"date={date}, exdate={exdate} (need >= 4, got {n_obs}). "
            f"Returning empty DataFrame.",
            UserWarning, stacklevel=2,
        )
        return _empty_result()

    # ── 2. Cubic spline on IV over moneyness ─────────────────────────────────
    spline = _fit_iv_spline(df_filt)
    if spline is None:
        return _empty_result()

    # ── 3. Uniform moneyness grid ────────────────────────────────────────────
    grid = _build_uniform_grid(df_filt, S, spline, n_grid, df_obs=df_filt)

    # ── 4. Reconstruct prices via Black (1976) ───────────────────────────────
    # Use puts for K <= F, calls for K > F (Bates 2005 / BKM 2003 convention)
    cp_flags = np.where(grid["strike"].values <= S, -1.0, 1.0)
    grid["cp_flag"]  = np.where(cp_flags < 0, "P", "C")
    grid["bs_price"] = _bs_option_price(
        K=grid["strike"].values, F=S, r=r, T=T,
        sigma=grid["iv_spline"].values, cp=cp_flags,
    )

    # ── 5 & 6. Numerical differentiation → Bates Greeks ─────────────────────
    return _compute_greeks(grid, S, date, exdate, security_id, T)


# ─── Private helpers ──────────────────────────────────────────────────────────

def _get_forward_price(
    source: Union[ForwardPrice, pd.DataFrame],
    date: str,
    exdate: str,
    security_id: int,
) -> float | None:
    """
    Retrieve the forward price for (date, exdate, security_id).

    Accepts either a ForwardPrice instance (reads ``.df``) or a plain DataFrame
    with columns ``[SecurityID, Date, Expiration, ForwardPrice]``.
    Returns ``None`` when no matching row is found.
    """
    df_fp: pd.DataFrame | None = source.df if hasattr(source, "df") else source
    if df_fp is None or len(df_fp) == 0:
        return None

    df_fp = df_fp.copy()
    df_fp["Date"]       = pd.to_datetime(df_fp["Date"],       errors="coerce")
    df_fp["Expiration"] = pd.to_datetime(df_fp["Expiration"], errors="coerce")

    mask = (
        (df_fp["SecurityID"] == security_id)          &
        (df_fp["Date"]       == pd.to_datetime(date)) &
        (df_fp["Expiration"] == pd.to_datetime(exdate))
    )
    subset = df_fp.loc[mask, "ForwardPrice"]
    return float(subset.iloc[0]) if len(subset) > 0 else None


def _filter_options(
    df: pd.DataFrame,
    date: str,
    exdate: str,
    security_id: int,
    S: float,
) -> tuple[pd.DataFrame | None, float, float]:
    """
    Filter OPPRCD data for (date, exdate, security_id) and apply OTM / quality cuts.

    Returns
    -------
    (df_filtered, r, T)
        ``df_filtered`` is ``None`` when no rows survive; ``r`` and ``T`` are
        derived from the first matching row before option-level filters.
    """
    df = df.copy()
    df["date"]   = pd.to_datetime(df["date"],   errors="coerce")
    df["exdate"] = pd.to_datetime(df["exdate"], errors="coerce")

    subset = df.loc[
        (df["date"]        == pd.to_datetime(date))   &
        (df["exdate"]      == pd.to_datetime(exdate)) &
        (df["security_id"] == security_id)
    ].copy()

    if len(subset) == 0:
        return None, 0.0, 0.0

    T = (pd.to_datetime(exdate) - pd.to_datetime(date)).days / 365.0
    if T <= 0:
        return None, 0.0, 0.0

    r_series = subset["tb_m3"].dropna()
    r = float(r_series.iloc[0]) if len(r_series) > 0 else 0.0

    # OTM: puts with K < S, calls with K > S
    mask_otm = (
        ((subset["cp_flag"] == "P") & (subset["strike_price"] <  S)) |
        ((subset["cp_flag"] == "C") & (subset["strike_price"] >  S))
    )
    subset = subset.loc[mask_otm].copy()

    # Quality filters
    subset = subset.loc[
        (subset["best_bid"]        >  0.0)    &
        (subset["best_offer"]      >  0.0)    &
        (subset["impl_volatility"] >  0.0)    &
        (subset["impl_volatility"] != -99.99)
    ].copy()

    if len(subset) == 0:
        return None, r, T

    subset["mid_price"] = (subset["best_bid"] + subset["best_offer"]) / 2.0
    subset["moneyness"] = subset["strike_price"] / S
    subset = subset.sort_values("moneyness").reset_index(drop=True)

    return subset, r, T


def _fit_iv_spline(df: pd.DataFrame) -> CubicSpline | None:
    """
    Fit a cubic spline on (moneyness, impl_volatility).

    Duplicate moneyness values are averaged. Requires >= 4 unique knots.
    Returns ``None`` and emits a warning if insufficient knots are available.
    """
    x = df["moneyness"].values
    y = df["impl_volatility"].values

    unique_x, inverse = np.unique(x, return_inverse=True)
    unique_y = np.array([y[inverse == i].mean() for i in range(len(unique_x))])

    if len(unique_x) < 4:
        warnings.warn(
            f"Cannot fit cubic spline: only {len(unique_x)} unique moneyness "
            f"values (need >= 4).",
            UserWarning, stacklevel=3,
        )
        return None

    return CubicSpline(unique_x, unique_y, extrapolate=False)


def _build_uniform_grid(
    df: pd.DataFrame,
    S: float,
    spline: CubicSpline,
    n: int,
    df_obs: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Evaluate the IV spline on a uniform moneyness grid of *n* points.

    Grid spans [min(moneyness), max(moneyness)] from observed data.
    Kj = mj * S;  DeltaK = (K_max - K_min) / (n - 1) is constant by construction.

    If ``df_obs`` is provided (filtered observed options), the observed mid-prices
    are linearly interpolated onto the grid.

    Returns
    -------
    pd.DataFrame with columns ``[moneyness, strike, iv_spline, mid_price]``.
    """
    m_min = df["moneyness"].min()
    m_max = df["moneyness"].max()

    m_grid  = np.linspace(m_min, m_max, n)
    iv_grid = np.asarray(spline(m_grid), dtype=float)
    iv_grid = np.clip(iv_grid, 1e-6, None)   # spline can produce tiny negatives near edges

    if df_obs is not None and "mid_price" in df_obs.columns:
        obs_sorted  = df_obs.sort_values("moneyness")
        mid_grid    = np.interp(
            m_grid,
            obs_sorted["moneyness"].values,
            obs_sorted["mid_price"].values,
        )
    else:
        mid_grid = np.full(n, np.nan)

    return pd.DataFrame({
        "moneyness": m_grid,
        "strike":    m_grid * S,
        "iv_spline": iv_grid,
        "mid_price": mid_grid,
    })


def _bs_option_price(
    K: np.ndarray,
    F: float,
    r: float,
    T: float,
    sigma: np.ndarray,
    cp: np.ndarray,
) -> np.ndarray:
    """
    European option price via Black (1976) using the forward price F.

    price = exp(-r*T) * phi * (F*N(phi*d1) - K*N(phi*d2))

    where phi = +1 for calls, -1 for puts, and
          d1 = [ln(F/K) + 0.5*sigma^2*T] / (sigma*sqrt(T))
          d2 = d1 - sigma*sqrt(T)

    Parameters
    ----------
    K     : array of strikes
    F     : forward price (OptionMetrics ForwardPrice, dividend-adjusted)
    r     : risk-free rate (annualized, decimal)
    T     : time to expiry in years
    sigma : array of implied volatilities
    cp    : +1.0 for calls, -1.0 for puts (same length as K / sigma)
    """
    K     = np.asarray(K,     dtype=float)
    sigma = np.asarray(sigma, dtype=float)
    cp    = np.asarray(cp,    dtype=float)
    sqrtT = np.sqrt(T)
    disc  = np.exp(-r * T)

    with np.errstate(divide="ignore", invalid="ignore"):
        d1 = (np.log(F / K) + 0.5 * sigma**2 * T) / (sigma * sqrtT)
        d2 = d1 - sigma * sqrtT

    return disc * cp * (F * norm.cdf(cp * d1) - K * norm.cdf(cp * d2))


def _numerical_derivatives(
    O: np.ndarray,
    dK: float,
) -> tuple[np.ndarray, np.ndarray]:
    """
    Central finite differences on a uniform-DeltaK grid.

    Only interior points i = 1, ..., n-2 are returned (boundary points dropped).

    Returns
    -------
    (O_X, O_XX) : arrays of length n-2.
    """
    O_X  = (O[2:] - O[:-2]) / (2.0 * dK)
    O_XX = (O[2:] - 2.0 * O[1:-1] + O[:-2]) / (dK**2)
    return O_X, O_XX


def _compute_greeks(
    grid: pd.DataFrame,
    S: float,
    date: str,
    exdate: str,
    security_id: int,
    T: float,
) -> pd.DataFrame:
    """
    Apply Bates (2005) equations (1) and (2).

    delta_bates[i] = (1/S) * (O[i] - K[i] * O_X[i])
    gamma_bates[i] = (K[i]/S)^2 * O_XX[i]

    Boundary points are discarded; only interior grid points are returned.
    """
    K  = grid["strike"].values
    O  = grid["bs_price"].values
    m  = grid["moneyness"].values
    iv = grid["iv_spline"].values

    n  = len(K)
    dK = (K[-1] - K[0]) / (n - 1)

    O_X, O_XX = _numerical_derivatives(O, dK)

    mp = grid["mid_price"].values
    cp = grid["cp_flag"].values

    K_int  = K[1:-1]
    O_int  = O[1:-1]
    m_int  = m[1:-1]
    iv_int = iv[1:-1]
    mp_int = mp[1:-1]
    cp_int = cp[1:-1]

    delta_bates = (1.0 / S) * (O_int - K_int * O_X)
    gamma_bates = (K_int / S)**2 * O_XX

    # Mask interior points whose finite-difference stencil crosses the put→call
    # boundary (left neighbor is put, right neighbor is call or vice versa).
    # At those points O_X mixes option types → derivative has wrong sign.
    stencil_crosses = cp[:-2] != cp[2:]   # length n-2, aligns with O_X / O_XX
    delta_bates = np.where(stencil_crosses, np.nan, delta_bates)
    gamma_bates = np.where(stencil_crosses, np.nan, gamma_bates)

    days_to_expiry = int(round(T * 365))

    # ── Column order: identifiers first, computed columns last ───────────────
    return pd.DataFrame({
        # --- original / identifier ---
        "date":           date,
        "exdate":         exdate,
        "security_id":    security_id,
        "strike":         K_int,
        "cp_flag":        cp_int,
        # --- computed (light-blue header in Excel) ---
        "days_to_expiry": days_to_expiry,
        "moneyness":      m_int,
        "mid_price":      mp_int,
        "iv_spline":      iv_int,
        "bs_price":       O_int,
        "delta_bates":    delta_bates,
        "gamma_bates":    gamma_bates,
    })


def _empty_result() -> pd.DataFrame:
    """Return an empty DataFrame with the canonical output columns."""
    return pd.DataFrame(columns=[
        "date", "exdate", "security_id", "strike", "cp_flag",
        "days_to_expiry", "moneyness", "mid_price", "iv_spline",
        "bs_price", "delta_bates", "gamma_bates",
    ])


# ─── Ejecución ────────────────────────────────────────────────────────────────

if __name__ == "__main__":

    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    from __2_Files.option_price  import OptionPrice
    from __2_Files.zero_curve    import ZeroCurve

    # ── Configuración ─────────────────────────────────────────────────────────

    PARQUET_DIR  = Path(r"C:\Users\pablo.esparcia\Documents\OptionMetrics\Acumulado")
    OUTPUT_PATH  = Path(r"C:\Users\pablo.esparcia\Downloads\bates_greeks_resultado.xlsx")

    YEARS        = [2015]
    SECURITY_ID  = 108105       # S&P 500
    CURRENCY     = 333          # USD
    N_GRID       = 100

    DESDE        = f"{min(YEARS)}-01-01"
    HASTA        = f"{max(YEARS)}-12-31"

    # ── 1. Cargar datos ───────────────────────────────────────────────────────

    print("Cargando OptionPrice...")
    op = OptionPrice()
    op.cargar_parquet(
        ruta     = str(PARQUET_DIR / "option_price.parquet"),
        desde    = DESDE,
        hasta    = HASTA,
        columnas = ["SecurityID", "Date", "Expiration", "Strike",
                    "CallPut", "Bid", "Ask", "ImpliedVolatility"],
    )

    print("Cargando ZeroCurve...")
    zc = ZeroCurve()
    zc.cargar_parquet(
        ruta  = str(PARQUET_DIR / "zero_curve.parquet"),
        desde = DESDE,
        hasta = HASTA,
    )

    print("Cargando ForwardPrice...")
    fp = ForwardPrice()
    fp.cargar_parquet(
        ruta     = str(PARQUET_DIR / "forward_price.parquet"),
        desde    = DESDE,
        hasta    = HASTA,
        columnas = ["SecurityID", "Date", "Expiration", "ForwardPrice"],
    )

    # ── 2. Tasa libre de riesgo a 90 días por fecha ───────────────────────────
    # ZeroCurve: Currency | Date | Days | Rate  (Days=90 ≈ tb_m3)
    # Se construye un mapa date → rate para el merge posterior.

    rates = (
        zc.df
        .loc[(zc.df["Currency"] == CURRENCY) & (zc.df["Days"] == 90), ["Date", "Rate"]]
        .rename(columns={"Rate": "tb_m3"})
        .drop_duplicates("Date")
    )

    # ── 3. Preparar DataFrame de opciones ─────────────────────────────────────

    df = op.df.rename(columns={
        "SecurityID":        "security_id",
        "Date":              "date",
        "Expiration":        "exdate",
        "Strike":            "strike_price",
        "CallPut":           "cp_flag",
        "Bid":               "best_bid",
        "Ask":               "best_offer",
        "ImpliedVolatility": "impl_volatility",
    }).copy()

    # Filtrar por security_id y años seleccionados
    df = df.loc[
        (df["security_id"] == SECURITY_ID) &
        (df["date"].dt.year.isin(YEARS))
    ].copy()

    # OptionMetrics raw TXT stores strikes ×1000 (e.g. 1500000 → $1500); adjust once.
    if df["strike_price"].median() > 10_000:
        print("AVISO: strikes detectados en unidades ×1000 → dividiendo por 1000.")
        df["strike_price"] = df["strike_price"] / 1000.0

    # Merge de la tasa por fecha
    df = df.merge(rates, left_on="date", right_on="Date", how="left").drop(columns="Date")
    df["tb_m3"] = df["tb_m3"].fillna(0.0)

    # ── 4. Pre-indexar ForwardPrice para lookup O(1) ──────────────────────────

    fp_df = fp.df.copy()
    fp_df["Date"]       = pd.to_datetime(fp_df["Date"],       errors="coerce")
    fp_df["Expiration"] = pd.to_datetime(fp_df["Expiration"], errors="coerce")
    fp_idx = (
        fp_df
        .groupby(["SecurityID", "Date", "Expiration"])["ForwardPrice"]
        .first()          # colapsa duplicados → escalar garantizado
        .sort_index()     # lexsort para lookup O(log n)
    )

    # ── 5. Pre-agrupar opciones por (date, exdate) ────────────────────────────

    grupos = {
        key: grp
        for key, grp in df.groupby(["date", "exdate"])
        if key[1] > key[0]   # descartar vencidos (T <= 0)
    }
    pares = sorted(grupos.keys())

    print(f"\n{len(pares):,} pares (date, exdate) a calcular para SecurityID={SECURITY_ID}...\n")

    # ── 6. Iterar ─────────────────────────────────────────────────────────────

    bloques = []
    for i, (date_ts, exdate_ts) in enumerate(pares):
        # Lookup O(1) del precio forward para este par
        fp_key = (SECURITY_ID, date_ts, exdate_ts)
        try:
            S_val = float(fp_idx.loc[fp_key])
        except KeyError:
            continue

        fp_slice = pd.DataFrame([{
            "SecurityID":   SECURITY_ID,
            "Date":         date_ts,
            "Expiration":   exdate_ts,
            "ForwardPrice": S_val,
        }])

        res = compute_bates_greeks(
            df_opprcd            = grupos[(date_ts, exdate_ts)],
            forward_price_source = fp_slice,
            date                 = date_ts.strftime("%Y-%m-%d"),
            exdate               = exdate_ts.strftime("%Y-%m-%d"),
            security_id          = SECURITY_ID,
            n_grid               = N_GRID,
        )
        if not res.empty:
            bloques.append(res)

        if (i + 1) % 500 == 0:
            print(f"  {i + 1:,} / {len(pares):,} completados...")

    # ── 7. Consolidar y guardar ───────────────────────────────────────────────

    if not bloques:
        print("Sin resultados.")
    else:   
        resultado = pd.concat(bloques, ignore_index=True)
        print(f"\nTotal filas: {len(resultado):,}  |  pares con resultado: {len(bloques):,} / {len(pares):,}")

        # Restaurar nombres originales de OptionMetrics para las columnas identificadoras
        resultado = resultado.rename(columns={
            "date":        "Date",
            "exdate":      "Expiration",
            "security_id": "SecurityID",
            "strike":      "Strike",
            "cp_flag":     "CallPut",
        })

        # Columnas originales vs. calculadas
        COLS_ORIGINAL  = ["Date", "Expiration", "SecurityID", "Strike", "CallPut"]
        COLS_COMPUTED  = [c for c in resultado.columns if c not in COLS_ORIGINAL]

        fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        font_bold = Font(bold=True)

        with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
            resultado.to_excel(writer, index=False, sheet_name="Greeks")
            ws = writer.sheets["Greeks"]

            for cell in ws[1]:                          # fila de cabeceros
                cell.font = font_bold
                if cell.value in COLS_COMPUTED:
                    cell.fill = fill_blue

            # Autoajuste de ancho de columnas
            for col_idx, col_name in enumerate(resultado.columns, start=1):
                width = max(len(str(col_name)), 10) + 2
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        print(f"Guardado en: {OUTPUT_PATH}")
